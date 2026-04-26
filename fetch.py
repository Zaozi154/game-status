import json
from openpyxl import load_workbook

# 读取仓库中的 Excel 文件
wb = load_workbook('data.xlsx', data_only=True)

# Sheet 顺序与之前一致（从左到右）
all_sheets = ['角色','技能','角色技能','伙伴','伙伴技能','角色伙伴','物品','装备配置']

sheets = {}
for name in all_sheets:
    if name not in wb.sheetnames:
        print(f"⚠️ 未找到 Sheet：{name}，跳过")
        continue
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print(f"⚠️ {name} 没有数据行，跳过")
        continue
    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        # 跳过完全空行
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        entry = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = str(cell).strip() if cell is not None else ''
            else:
                entry[f'col{i}'] = str(cell).strip() if cell is not None else ''
        data.append(entry)
    sheets[name] = data
    print(f"  {name}: {len(data)} 条")

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(sheets, f, ensure_ascii=False, indent=2)

print("✅ data.json 已从 data.xlsx 生成")
